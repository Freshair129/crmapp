"""
generate_master_data_template.py
Generates vschool_master_data_template.xlsx — Google Sheets master data template
for V School CRM sync API (ADR-036).
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

OUTPUT_PATH = "/Users/ideab/Desktop/crm/vschool_master_data_template.xlsx"

# ── Colour palette ────────────────────────────────────────────────────────────
COLORS = {
    "courses_header":     "1F4E79",
    "ingredients_header": "1B5E20",
    "bom_header":         "4A148C",
    "assets_header":      "B71C1C",
    "white":              "FFFFFF",
    "note_bg":            "FFFDE7",   # light yellow
    "alt_row":            "F5F5F5",
    "instr_title":        "1F4E79",
    "instr_step_bg":      "E3F2FD",
    "table_header_bg":    "CFE2F3",
}

# ── Style helpers ─────────────────────────────────────────────────────────────

def header_font(size=10):
    return Font(name="Arial", size=size, bold=True, color=COLORS["white"])

def data_font(size=10):
    return Font(name="Arial", size=size)

def make_fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def thin_border():
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def apply_header_row(ws, headers, col_widths, header_color, row=1):
    """Style the header row with background colour, white bold text, borders."""
    fill = make_fill(header_color)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = header_font()
        cell.fill = fill
        cell.alignment = center()
        cell.border = thin_border()
    ws.row_dimensions[row].height = 28

    # Column widths
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def apply_data_row(ws, row_idx, values, alt=False):
    """Write a data row with alternating background, borders, and Arial 10."""
    fill = make_fill(COLORS["alt_row"]) if alt else PatternFill()
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = data_font()
        cell.alignment = left()
        cell.border = thin_border()
        if alt:
            cell.fill = fill

def add_note_row(ws, row_idx, text, num_cols):
    """Merge a full-width note row with light-yellow background."""
    ws.merge_cells(
        start_row=row_idx, start_column=1,
        end_row=row_idx, end_column=num_cols
    )
    cell = ws.cell(row=row_idx, column=1, value=text)
    cell.font = Font(name="Arial", size=9, italic=True, color="5D4037")
    cell.fill = make_fill(COLORS["note_bg"])
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = thin_border()
    ws.row_dimensions[row_idx].height = 22

def freeze_and_set_data_heights(ws, data_rows):
    ws.freeze_panes = "A2"
    for row_idx in range(2, 2 + data_rows):
        ws.row_dimensions[row_idx].height = 20


# ── Sheet 1: คอร์ส (Courses) ──────────────────────────────────────────────────

def build_courses(wb):
    ws = wb.create_sheet("คอร์ส (Courses)")
    headers = ["productId", "name", "category", "price", "duration", "description"]
    widths  = [20, 35, 15, 12, 12, 40]

    apply_header_row(ws, headers, widths, COLORS["courses_header"])

    data = [
        ("PRD-CRS-2026-001", "สูตรซูชิพรีเมียม",      "course", 3500, 6, "เรียนทำซูชิสไตล์ญี่ปุ่นแท้"),
        ("PRD-CRS-2026-002", "ราเมนต้นตำรับ",         "course", 2800, 4, "ราเมนน้ำซุปโชยุและมิโซะ"),
        ("PRD-CRS-2026-003", "เทมปุระและคาราอาเกะ",   "course", 2200, 3, "เทคนิคทอดญี่ปุ่น"),
    ]
    for i, row in enumerate(data, start=2):
        apply_data_row(ws, i, row, alt=(i % 2 == 1))

    # blank row 5, note row 6
    ws.row_dimensions[5].height = 6
    add_note_row(
        ws, 6,
        "⚠️ หมายเหตุ: productId ต้องไม่ซ้ำ • category: course / package • duration = ชั่วโมง",
        len(headers)
    )

    freeze_and_set_data_heights(ws, len(data))


# ── Sheet 2: วัตถุดิบ (Ingredients) ──────────────────────────────────────────

def build_ingredients(wb):
    ws = wb.create_sheet("วัตถุดิบ (Ingredients)")
    headers = ["ingredientId", "name", "unit", "currentStock", "minStock", "category", "costPerUnit"]
    widths  = [22, 30, 12, 15, 12, 18, 14]

    apply_header_row(ws, headers, widths, COLORS["ingredients_header"])

    data = [
        ("ING-2026-001", "ข้าวญี่ปุ่น",          "กก.",   50,  10, "grain",   85),
        ("ING-2026-002", "ปลาทูน่าสดแช่แข็ง",    "กก.",    5,   2, "seafood", 450),
        ("ING-2026-003", "สาหร่ายโนริ",           "แผ่น", 200,  50, "dry",      8),
        ("ING-2026-004", "ซอสโชยุ",               "ลิตร",  10,   3, "sauce",  120),
    ]
    for i, row in enumerate(data, start=2):
        apply_data_row(ws, i, row, alt=(i % 2 == 1))

    ws.row_dimensions[6].height = 6
    add_note_row(
        ws, 7,
        "⚠️ หมายเหตุ: currentStock = สต็อกปัจจุบัน • minStock = จุดสั่งซื้อ • costPerUnit = ราคาต่อหน่วย (บาท)",
        len(headers)
    )

    freeze_and_set_data_heights(ws, len(data))


# ── Sheet 3: BOM ──────────────────────────────────────────────────────────────

def build_bom(wb):
    ws = wb.create_sheet("BOM (สูตร-วัตถุดิบต่อคน)")
    headers = ["productId", "ingredientId", "qtyPerPerson", "unit"]
    widths  = [25, 22, 16, 12]

    apply_header_row(ws, headers, widths, COLORS["bom_header"])

    data = [
        ("PRD-CRS-2026-001", "ING-2026-001", 0.15, "กก."),
        ("PRD-CRS-2026-001", "ING-2026-002", 0.08, "กก."),
        ("PRD-CRS-2026-001", "ING-2026-003", 3,    "แผ่น"),
        ("PRD-CRS-2026-002", "ING-2026-001", 0.2,  "กก."),
        ("PRD-CRS-2026-002", "ING-2026-004", 0.05, "ลิตร"),
    ]
    for i, row in enumerate(data, start=2):
        apply_data_row(ws, i, row, alt=(i % 2 == 1))

    ws.row_dimensions[8].height = 6
    add_note_row(
        ws, 9,
        "⚠️ หมายเหตุ: productId และ ingredientId ต้องมีอยู่ในชีต คอร์ส และ วัตถุดิบ ก่อน • qtyPerPerson = ปริมาณต่อนักเรียน 1 คน",
        len(headers)
    )

    freeze_and_set_data_heights(ws, len(data))


# ── Sheet 4: อุปกรณ์ (Assets) ────────────────────────────────────────────────

def build_assets(wb):
    ws = wb.create_sheet("อุปกรณ์ (Assets)")
    headers = ["assetId", "name", "category", "status", "location",
               "purchasePrice", "vendor", "serialNumber", "notes"]
    widths  = [22, 30, 15, 12, 18, 15, 20, 20, 35]

    apply_header_row(ws, headers, widths, COLORS["assets_header"])

    data = [
        ("AST-KIT-2026-001", "มีดซูชิมืออาชีพ",     "KITCHEN",   "ACTIVE", "ครัวห้อง A",   2500, "Yoshida Knives",  "YK-2026-001", "คมมาก ระวังการใช้งาน"),
        ("AST-KIT-2026-002", "เขียงไม้ฮิโนกิ",       "KITCHEN",   "ACTIVE", "ครัวห้อง A",   1800, "Hinoki Wood Co.", "HW-001",      ""),
        ("AST-EQP-2026-001", "โปรเจกเตอร์ Epson",   "EQUIPMENT", "ACTIVE", "ห้องเรียน 1", 25000, "Epson Thailand",  "EPS-X2026",   "ใช้สำหรับสอนเทคนิค"),
        ("AST-EQP-2026-002", "กระดานไวท์บอร์ด",     "EQUIPMENT", "ACTIVE", "ห้องเรียน 1",  3500, "Office Mate",     "",            ""),
    ]
    for i, row in enumerate(data, start=2):
        apply_data_row(ws, i, row, alt=(i % 2 == 1))

    ws.row_dimensions[6].height = 6
    add_note_row(
        ws, 7,
        "⚠️ หมายเหตุ: category: KITCHEN / EQUIPMENT / VEHICLE / BUILDING / IT • status: ACTIVE / MAINTENANCE / RETIRED",
        len(headers)
    )

    freeze_and_set_data_heights(ws, len(data))


# ── Sheet 5: วิธีใช้ (Instructions) ──────────────────────────────────────────

def build_instructions(wb):
    ws = wb.create_sheet("วิธีใช้ (Instructions)")
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 35

    # Title
    ws.merge_cells("A1:C1")
    title = ws["A1"]
    title.value = "คู่มือการใช้งาน Google Sheets Master Data"
    title.font = Font(name="Arial", size=14, bold=True, color=COLORS["instr_title"])
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 36

    # Subtitle / version
    ws.merge_cells("A2:C2")
    sub = ws["A2"]
    sub.value = "V School CRM v2 — ADR-036 · sync API: /api/sheets/sync-master-data"
    sub.font = Font(name="Arial", size=9, italic=True, color="757575")
    sub.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 10  # spacer

    # Steps
    steps = [
        ("1.", "กรอกข้อมูลแต่ละหมวดหมู่ในชีตที่กำหนด อย่าเปลี่ยนชื่อ header row"),
        ("2.", "แชร์ Google Sheet แบบ \"ทุกคนที่มีลิงก์สามารถดู\""),
        ("3.", "คัดลอก URL ของแต่ละชีต ในรูปแบบ CSV:\n   เปลี่ยน /edit#gid=xxx  →  /export?format=csv&gid=xxx"),
        ("4.", "วาง URL ใน .env:\n   SHEET_COURSES_URL, SHEET_INGREDIENTS_URL, SHEET_BOM_URL, SHEET_ASSETS_URL"),
        ("5.", "ใช้ปุ่ม Sync ใน CRM เพื่ออัปเดตข้อมูล"),
    ]
    step_fill = make_fill(COLORS["instr_step_bg"])
    for row_offset, (num, text) in enumerate(steps):
        r = 4 + row_offset
        ws.row_dimensions[r].height = 32
        n_cell = ws.cell(row=r, column=1, value=num)
        n_cell.font = Font(name="Arial", size=10, bold=True, color=COLORS["instr_title"])
        n_cell.alignment = center()
        n_cell.fill = step_fill
        n_cell.border = thin_border()

        t_cell = ws.cell(row=r, column=2, value=text)
        t_cell.font = data_font()
        t_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        t_cell.fill = step_fill
        t_cell.border = thin_border()

    ws.row_dimensions[9].height = 14  # spacer

    # Table: .env key → sheet name
    ws.merge_cells("A10:C10")
    th = ws["A10"]
    th.value = "ตาราง: .env key → ชีต"
    th.font = Font(name="Arial", size=10, bold=True, color=COLORS["white"])
    th.fill = make_fill(COLORS["instr_title"])
    th.alignment = center()
    th.border = thin_border()
    ws.row_dimensions[10].height = 24

    table_header_fill = make_fill(COLORS["table_header_bg"])
    env_headers = [("", "Environment Variable (.env)", "Sheet Name")]
    env_rows = [
        ("→", "SHEET_COURSES_URL",     "คอร์ส (Courses)"),
        ("→", "SHEET_INGREDIENTS_URL", "วัตถุดิบ (Ingredients)"),
        ("→", "SHEET_BOM_URL",         "BOM (สูตร/วัตถุดิบต่อคน)"),
        ("→", "SHEET_ASSETS_URL",      "อุปกรณ์ (Assets)"),
    ]

    # Table header
    for col_idx, val in enumerate(env_headers[0], start=1):
        cell = ws.cell(row=11, column=col_idx, value=val)
        cell.font = Font(name="Arial", size=10, bold=True)
        cell.fill = table_header_fill
        cell.alignment = center()
        cell.border = thin_border()
    ws.row_dimensions[11].height = 22

    for r_off, row in enumerate(env_rows):
        r = 12 + r_off
        alt = r_off % 2 == 1
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.font = data_font()
            cell.alignment = Alignment(horizontal="left" if col_idx > 1 else "center",
                                       vertical="center")
            cell.border = thin_border()
            if alt:
                cell.fill = make_fill(COLORS["alt_row"])
        ws.row_dimensions[r].height = 20

    # Footer note
    ws.row_dimensions[17].height = 10
    ws.merge_cells("A18:C18")
    footer = ws["A18"]
    footer.value = "สร้างโดย generate_master_data_template.py · V School CRM v2 · ADR-036"
    footer.font = Font(name="Arial", size=8, italic=True, color="9E9E9E")
    footer.alignment = Alignment(horizontal="right", vertical="center")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    build_courses(wb)
    build_ingredients(wb)
    build_bom(wb)
    build_assets(wb)
    build_instructions(wb)

    wb.save(OUTPUT_PATH)
    print(f"Saved: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
