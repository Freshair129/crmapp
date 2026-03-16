#!/usr/bin/env python3
"""
Generate vschool_db_import.xlsx — V School CRM data-entry workbook
Full PostgreSQL schema mirror with dropdowns and FK references.
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    GradientFill
)
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter

OUTPUT_PATH = "/Users/ideab/Desktop/crm/vschool_db_import.xlsx"

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def header_fill(hex_color):
    return PatternFill('solid', start_color=hex_color.lstrip('#'), end_color=hex_color.lstrip('#'))

REQUIRED_FILL = PatternFill('solid', start_color='FFFDE7', end_color='FFFDE7')
FK_FILL       = PatternFill('solid', start_color='E3F2FD', end_color='E3F2FD')
ALT_FILL      = PatternFill('solid', start_color='F5F5F5', end_color='F5F5F5')
WHITE_FILL    = PatternFill('solid', start_color='FFFFFF', end_color='FFFFFF')

FONT_NORMAL   = Font(name='Arial', size=10)
FONT_HEADER   = Font(name='Arial', size=10, bold=True, color='FFFFFF')

BORDER = thin_border()
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT   = Alignment(horizontal='left', vertical='center')


def style_header_cell(cell, fill, required=False, is_fk=False):
    """Style a header cell with domain colour; required/FK overrides tint the fill differently."""
    cell.font = FONT_HEADER
    cell.fill = fill
    cell.alignment = CENTER
    cell.border = BORDER


def style_data_cell(cell, row_idx):
    cell.font = FONT_NORMAL
    cell.fill = WHITE_FILL if row_idx % 2 == 0 else ALT_FILL
    cell.alignment = LEFT
    cell.border = BORDER


def write_headers(ws, columns, domain_fill, required_cols=(), fk_cols=()):
    """
    columns: list of (header_label, col_letter)  OR just a list of labels (A, B, C…)
    We accept list of label strings; col letters derived by position.
    required_cols / fk_cols: sets of header labels
    """
    ws.row_dimensions[1].height = 28
    for i, label in enumerate(columns):
        col_letter = get_column_letter(i + 1)
        cell = ws[f"{col_letter}1"]
        cell.value = label
        # Base header style
        cell.font = FONT_HEADER
        cell.fill = domain_fill
        cell.alignment = CENTER
        cell.border = BORDER
        # Override fill for required / FK (subtle — keep text white but note in comment)
        # Per spec: required = yellow fill, FK = blue fill on header
        if label.rstrip('*') in [r.rstrip('*') for r in required_cols] or label in required_cols:
            cell.fill = REQUIRED_FILL
            cell.font = Font(name='Arial', size=10, bold=True, color='555500')
        if label.rstrip('*') in [f.rstrip('*') for f in fk_cols] or label in fk_cols:
            cell.fill = FK_FILL
            cell.font = Font(name='Arial', size=10, bold=True, color='003366')


def allocate_data_rows(ws, num_cols, start_row=2, end_row=201):
    """Pre-allocate 200 data rows with alternating fills and borders."""
    for r in range(start_row, end_row + 1):
        ws.row_dimensions[r].height = 18
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = FONT_NORMAL
            cell.fill = WHITE_FILL if r % 2 == 0 else ALT_FILL
            cell.alignment = LEFT
            cell.border = BORDER


def add_dropdown(ws, col_letter, formula1, start_row=2, end_row=201):
    dv = DataValidation(
        type="list",
        formula1=formula1,
        allow_blank=True,
        showDropDown=False
    )
    ws.add_data_validation(dv)
    dv.sqref = f"{col_letter}{start_row}:{col_letter}{end_row}"


def write_sample_row(ws, row_num, values, num_cols):
    """Write sample data into a row; style cells."""
    for c, val in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=c)
        cell.value = val
        cell.font = FONT_NORMAL
        cell.fill = WHITE_FILL if row_num % 2 == 0 else ALT_FILL
        cell.alignment = LEFT
        cell.border = BORDER


# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------

wb = Workbook()

# Remove default sheet
if 'Sheet' in wb.sheetnames:
    del wb['Sheet']

# ---------------------------------------------------------------------------
# 1. _Enums sheet
# ---------------------------------------------------------------------------

ws_enums = wb.create_sheet('_Enums')

ENUM_DATA = {
    'A': ('customer_status',   ['Active', 'Inactive', 'Blocked']),
    'B': ('membership_tier',   ['MEMBER', 'SILVER', 'GOLD', 'PLATINUM', 'VIP', 'DIAMOND']),
    'C': ('lifecycle_stage',   ['Lead', 'Prospect', 'Active', 'Customer', 'Churned']),
    'D': ('employee_role',     ['Developer', 'Manager', 'Supervisor', 'Admin', 'Agent', 'Guest']),
    'E': ('employee_status',   ['ACTIVE', 'INACTIVE', 'SUSPENDED']),
    'F': ('employee_dept',     ['MARKETING', 'KITCHEN', 'ADMIN', 'SALES', 'MANAGEMENT']),
    'G': ('asset_category',    ['KITCHEN', 'EQUIPMENT', 'VEHICLE', 'BUILDING', 'IT', 'GENERAL']),
    'H': ('asset_status',      ['ACTIVE', 'MAINTENANCE', 'RETIRED']),
    'I': ('ingredient_category',['grain', 'seafood', 'dry', 'sauce', 'vegetable', 'meat', 'dairy', 'other']),
    'J': ('unit_list',         ['กก.', 'กรัม', 'ลิตร', 'มล.', 'แผ่น', 'ชิ้น', 'อัน', 'ถ้วย', 'ลัง', 'แพ็ก', 'kg', 'L']),
    'K': ('product_category',  ['course', 'package']),
    'L': ('session_type',      ['MORNING', 'AFTERNOON', 'EVENING']),
    'M': ('schedule_status',   ['OPEN', 'FULL', 'CANCELLED', 'COMPLETED']),
    'N': ('enrollment_status', ['ACTIVE', 'COMPLETED', 'CANCELLED']),
    'O': ('recipe_category',   ['JP', 'TH', 'PASTRY', 'WESTERN', 'CHINESE']),
    'P': ('recipe_chef',       ['AOI', 'FAH', 'BKK']),
    'Q': ('task_type',         ['FOLLOW_UP', 'CALL', 'EMAIL', 'MEETING', 'DEMO']),
    'R': ('task_priority',     ['LOW', 'MEDIUM', 'HIGH', 'URGENT']),
    'S': ('task_status',       ['PENDING', 'IN_PROGRESS', 'COMPLETED', 'CANCELLED']),
    'T': ('pr_status',         ['DRAFT', 'PENDING_APPROVAL', 'APPROVED', 'ORDERED', 'RECEIVED', 'CANCELLED']),
    'U': ('bool_field',        ['TRUE', 'FALSE']),
    'V': ('cert_level',        ['30', '111', '201']),
    'W': ('order_status',      ['PENDING', 'CONFIRMED', 'COMPLETED', 'CANCELLED']),
}

for col_letter, (header, values) in ENUM_DATA.items():
    ws_enums[f'{col_letter}1'] = header
    ws_enums[f'{col_letter}1'].font = Font(name='Arial', size=10, bold=True)
    for i, val in enumerate(values, start=2):
        ws_enums[f'{col_letter}{i}'] = val
        ws_enums[f'{col_letter}{i}'].font = Font(name='Arial', size=10)

# Named ranges — exact cell refs from spec
NAMED_RANGES = {
    'customer_status':    '_Enums!$A$2:$A$4',
    'membership_tier':    '_Enums!$B$2:$B$7',
    'lifecycle_stage':    '_Enums!$C$2:$C$6',
    'employee_role':      '_Enums!$D$2:$D$7',
    'employee_status':    '_Enums!$E$2:$E$4',
    'employee_dept':      '_Enums!$F$2:$F$6',
    'asset_category':     '_Enums!$G$2:$G$7',
    'asset_status':       '_Enums!$H$2:$H$4',
    'ingredient_category':'_Enums!$I$2:$I$9',
    'unit_list':          '_Enums!$J$2:$J$13',
    'product_category':   '_Enums!$K$2:$K$3',
    'session_type':       '_Enums!$L$2:$L$4',
    'schedule_status':    '_Enums!$M$2:$M$5',
    'enrollment_status':  '_Enums!$N$2:$N$4',
    'recipe_category':    '_Enums!$O$2:$O$6',
    'recipe_chef':        '_Enums!$P$2:$P$4',
    'task_type':          '_Enums!$Q$2:$Q$6',
    'task_priority':      '_Enums!$R$2:$R$5',
    'task_status':        '_Enums!$S$2:$S$5',
    'pr_status':          '_Enums!$T$2:$T$7',
    'bool_field':         '_Enums!$U$2:$U$3',
    'cert_level':         '_Enums!$V$2:$V$4',
    'order_status':       '_Enums!$W$2:$W$5',
}

for name, attr_text in NAMED_RANGES.items():
    wb.defined_names[name] = DefinedName(name, attr_text=attr_text)

ws_enums.sheet_state = 'hidden'

# ---------------------------------------------------------------------------
# 2. คู่มือ sheet
# ---------------------------------------------------------------------------

ws_guide = wb.create_sheet('คู่มือ')

# Title
title_cell = ws_guide['A1']
title_cell.value = 'คู่มือการนำเข้าข้อมูล V School CRM'
title_cell.font = Font(name='Arial', size=16, bold=True, color='1F4E79')
title_cell.alignment = Alignment(horizontal='left', vertical='center')
ws_guide.row_dimensions[1].height = 30

sub_cell = ws_guide['A2']
sub_cell.value = 'กรอกข้อมูลในแต่ละชีต แล้วส่งไฟล์ให้ Admin อัพโหลดเข้าระบบ'
sub_cell.font = Font(name='Arial', size=11, color='555555')
ws_guide.row_dimensions[2].height = 22

ws_guide.row_dimensions[3].height = 10  # spacer

# Section: วิธีใช้งาน
sec1 = ws_guide['A4']
sec1.value = 'วิธีใช้งาน'
sec1.font = Font(name='Arial', size=13, bold=True, color='1F4E79')
ws_guide.row_dimensions[4].height = 24

instructions = [
    '1. ชีตที่ * = จำเป็นต้องกรอก (required)',
    '2. ช่องสีเหลือง = required field',
    '3. ช่องสีฟ้า = ต้องเลือกจาก dropdown (FK reference ไปยังชีตอื่น)',
    '4. อย่าเปลี่ยนชื่อ header row หรือชื่อชีต เพราะระบบอ่านตาม column name',
    '5. กรอกข้อมูลตั้งแต่ row 2 ลงไป',
    '6. วันที่ใช้รูปแบบ YYYY-MM-DD เช่น 2026-03-16',
    '7. ราคาเป็นหน่วยบาท ไม่ต้องใส่ comma',
]

for i, text in enumerate(instructions, start=5):
    cell = ws_guide.cell(row=i, column=1)
    cell.value = text
    cell.font = Font(name='Arial', size=11)
    cell.alignment = Alignment(horizontal='left', vertical='center')
    ws_guide.row_dimensions[i].height = 20

ws_guide.row_dimensions[5 + len(instructions)].height = 10  # spacer

# Section: ชีตและความสัมพันธ์
sec2_row = 5 + len(instructions) + 1
sec2 = ws_guide.cell(row=sec2_row, column=1)
sec2.value = 'ชีตและความสัมพันธ์'
sec2.font = Font(name='Arial', size=13, bold=True, color='1F4E79')
ws_guide.row_dimensions[sec2_row].height = 24

# Table headers
table_headers = ['ชีต', 'ID Format', 'FK ที่ใช้']
header_row = sec2_row + 1
for c, h in enumerate(table_headers, start=1):
    cell = ws_guide.cell(row=header_row, column=c)
    cell.value = h
    cell.font = Font(name='Arial', size=11, bold=True)
    cell.fill = PatternFill('solid', start_color='1F4E79', end_color='1F4E79')
    cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    cell.alignment = CENTER
    cell.border = BORDER
    ws_guide.row_dimensions[header_row].height = 22

sheet_info = [
    ('Customers',      'TVS-CUS-FB-26-XXXX',  '—'),
    ('Products',       'PRD-CRS-2026-XXX',     '—'),
    ('Employees',      'TVS-EMP-2026-XXXX',    '—'),
    ('CourseSchedules','SCH-2026-XXX',          'Products, Employees'),
    ('Enrollments',    'ENR-2026-XXXX',         'Customers, Products, Employees'),
    ('Ingredients',    'ING-2026-XXX',          '—'),
    ('CourseBOM',      '—',                     'Products, Ingredients'),
    ('Recipes',        'RCP-AOI-JP-XXX',        '—'),
    ('Packages',       'PKG-2026-XXX',          '—'),
    ('Assets',         'AST-KIT-2026-XXX',      'Employees'),
    ('Tasks',          'TSK-YYYYMMDD-XXX',      'Customers, Employees'),
]

for r_offset, (sheet_name, id_fmt, fk) in enumerate(sheet_info, start=1):
    row = header_row + r_offset
    ws_guide.row_dimensions[row].height = 20
    for c, val in enumerate([sheet_name, id_fmt, fk], start=1):
        cell = ws_guide.cell(row=row, column=c)
        cell.value = val
        cell.font = Font(name='Arial', size=11)
        cell.fill = WHITE_FILL if r_offset % 2 != 0 else ALT_FILL
        cell.alignment = LEFT
        cell.border = BORDER

ws_guide.column_dimensions['A'].width = 22
ws_guide.column_dimensions['B'].width = 25
ws_guide.column_dimensions['C'].width = 38

# ---------------------------------------------------------------------------
# 3. Customers sheet
# ---------------------------------------------------------------------------

ws_cus = wb.create_sheet('Customers')
DOMAIN_COLOR_CUS = '1565C0'
FILL_CUS = PatternFill('solid', start_color=DOMAIN_COLOR_CUS, end_color=DOMAIN_COLOR_CUS)

CUS_HEADERS = [
    'customerId*', 'memberId', 'firstName', 'lastName', 'nickName',
    'email', 'phonePrimary', 'lineId', 'facebookName',
    'status*', 'membershipTier*', 'lifecycleStage*',
    'walletBalance', 'walletPoints', 'jobTitle', 'company',
    'joinDate', 'notes (intelligence)'
]

CUS_REQUIRED = {'customerId*', 'status*', 'membershipTier*', 'lifecycleStage*'}
CUS_FK = set()
CUS_DROPDOWN = {
    'J': 'customer_status',
    'K': 'membership_tier',
    'L': 'lifecycle_stage',
}

ws_cus.freeze_panes = 'A2'
write_headers(ws_cus, CUS_HEADERS, FILL_CUS, CUS_REQUIRED, CUS_FK)
allocate_data_rows(ws_cus, len(CUS_HEADERS))

for col_letter, named_range in CUS_DROPDOWN.items():
    add_dropdown(ws_cus, col_letter, named_range)

# Sample row
sample_cus = [
    'TVS-CUS-FB-26-0001', 'MEM-26BKKP-0001', 'สมชาย', 'ใจดี', 'ชาย',
    'somchai@email.com', '+66812345678', 'line_somchai', 'Somchai Jaidee',
    'Active', 'GOLD', 'Active', 0, 0, 'Chef', '', '2026-01-15', ''
]
write_sample_row(ws_cus, 2, sample_cus, len(CUS_HEADERS))

# Column widths
col_widths_cus = [22, 20, 14, 14, 12, 24, 18, 16, 22, 12, 16, 14, 14, 14, 16, 16, 14, 28]
for i, w in enumerate(col_widths_cus, start=1):
    ws_cus.column_dimensions[get_column_letter(i)].width = w

# ---------------------------------------------------------------------------
# 4. Products sheet
# ---------------------------------------------------------------------------

ws_prd = wb.create_sheet('Products')
FILL_PRD = PatternFill('solid', start_color='1B5E20', end_color='1B5E20')

PRD_HEADERS = [
    'productId*', 'name*', 'category*', 'price*', 'duration',
    'days', 'sessionType', 'hours', 'description', 'isActive'
]
PRD_REQUIRED = {'productId*', 'name*', 'category*', 'price*'}
PRD_DROPDOWN = {
    'C': 'product_category',
    'G': 'session_type',
    'J': 'bool_field',
}

ws_prd.freeze_panes = 'A2'
write_headers(ws_prd, PRD_HEADERS, FILL_PRD, PRD_REQUIRED)
allocate_data_rows(ws_prd, len(PRD_HEADERS))

for col_letter, nr in PRD_DROPDOWN.items():
    add_dropdown(ws_prd, col_letter, nr)

samples_prd = [
    ['PRD-CRS-2026-001', 'สูตรซูชิพรีเมียม', 'course', 3500, 6, 1, 'MORNING', 6, 'เรียนทำซูชิสไตล์ญี่ปุ่นแท้', 'TRUE'],
    ['PRD-CRS-2026-002', 'ราเมนต้นตำรับ', 'course', 2800, 4, 1, 'AFTERNOON', 4, 'ราเมนน้ำซุปโชยุและมิโซะ', 'TRUE'],
    ['PRD-PKG-2026-001', 'แพ็กเกจซูชิ+ราเมน', 'package', 5800, '', 2, '', '', 'แพ็กเกจเรียน 2 คอร์ส', 'TRUE'],
]
for i, row_data in enumerate(samples_prd, start=2):
    write_sample_row(ws_prd, i, row_data, len(PRD_HEADERS))

col_widths_prd = [22, 24, 14, 10, 10, 8, 14, 8, 30, 10]
for i, w in enumerate(col_widths_prd, start=1):
    ws_prd.column_dimensions[get_column_letter(i)].width = w

# ---------------------------------------------------------------------------
# 5. Employees sheet
# ---------------------------------------------------------------------------

ws_emp = wb.create_sheet('Employees')
FILL_EMP = PatternFill('solid', start_color='4A148C', end_color='4A148C')

EMP_HEADERS = [
    'employeeId*', 'firstName*', 'lastName*', 'nickName', 'email*',
    'phone', 'department', 'role*', 'status*', 'passwordHash'
]
EMP_REQUIRED = {'employeeId*', 'firstName*', 'lastName*', 'email*', 'role*', 'status*'}
EMP_DROPDOWN = {
    'G': 'employee_dept',
    'H': 'employee_role',
    'I': 'employee_status',
}

ws_emp.freeze_panes = 'A2'
write_headers(ws_emp, EMP_HEADERS, FILL_EMP, EMP_REQUIRED)
allocate_data_rows(ws_emp, len(EMP_HEADERS))

for col_letter, nr in EMP_DROPDOWN.items():
    add_dropdown(ws_emp, col_letter, nr)

samples_emp = [
    ['TVS-EMP-2026-0001', 'Admin', 'User', 'Admin', 'admin@vschool.com', '', 'ADMIN', 'Admin', 'ACTIVE', ''],
    ['TVS-EMP-2026-0003', 'Fafah', 'Fasai', 'FAH', 'fafah@vschool.com', '', 'MARKETING', 'Agent', 'ACTIVE', ''],
    ['TVS-EMP-2026-0004', 'Satabongkot', 'Noinin', 'AOI', 'aoi@vschool.com', '', 'KITCHEN', 'Agent', 'ACTIVE', ''],
]
for i, row_data in enumerate(samples_emp, start=2):
    write_sample_row(ws_emp, i, row_data, len(EMP_HEADERS))

col_widths_emp = [22, 16, 16, 12, 26, 16, 16, 14, 12, 40]
for i, w in enumerate(col_widths_emp, start=1):
    ws_emp.column_dimensions[get_column_letter(i)].width = w

# ---------------------------------------------------------------------------
# 6. CourseSchedules sheet
# ---------------------------------------------------------------------------

ws_sch = wb.create_sheet('CourseSchedules')
FILL_SCH = PatternFill('solid', start_color='E65100', end_color='E65100')

SCH_HEADERS = [
    'scheduleId*', 'productId*', 'scheduledDate*', 'startTime', 'endTime',
    'sessionType', 'maxStudents', 'confirmedStudents', 'status*',
    'instructorId', 'notes'
]
SCH_REQUIRED = {'scheduleId*', 'productId*', 'scheduledDate*', 'status*'}
SCH_DROPDOWN_NAMED = {
    'F': 'session_type',
    'I': 'schedule_status',
}
SCH_DROPDOWN_FK = {
    'B': "'Products'!$A$2:$A$201",
    'J': "'Employees'!$A$2:$A$201",
}

ws_sch.freeze_panes = 'A2'
write_headers(ws_sch, SCH_HEADERS, FILL_SCH, SCH_REQUIRED)
allocate_data_rows(ws_sch, len(SCH_HEADERS))

for col_letter, nr in SCH_DROPDOWN_NAMED.items():
    add_dropdown(ws_sch, col_letter, nr)
for col_letter, formula in SCH_DROPDOWN_FK.items():
    add_dropdown(ws_sch, col_letter, formula)

sample_sch = ['SCH-2026-001', 'PRD-CRS-2026-001', '2026-04-01', '09:00', '17:00', 'MORNING', 10, 0, 'OPEN', 'TVS-EMP-2026-0004', '']
write_sample_row(ws_sch, 2, sample_sch, len(SCH_HEADERS))

col_widths_sch = [18, 22, 16, 12, 12, 14, 14, 18, 14, 22, 24]
for i, w in enumerate(col_widths_sch, start=1):
    ws_sch.column_dimensions[get_column_letter(i)].width = w

# ---------------------------------------------------------------------------
# 7. Enrollments sheet
# ---------------------------------------------------------------------------

ws_enr = wb.create_sheet('Enrollments')
FILL_ENR = PatternFill('solid', start_color='880E4F', end_color='880E4F')

ENR_HEADERS = [
    'enrollmentId*', 'customerId*', 'productId*', 'soldById',
    'totalPrice*', 'status*', 'enrolledAt', 'notes'
]
ENR_REQUIRED = {'enrollmentId*', 'customerId*', 'productId*', 'totalPrice*', 'status*'}
ENR_DROPDOWN_NAMED = {
    'F': 'enrollment_status',
}
ENR_DROPDOWN_FK = {
    'B': "'Customers'!$A$2:$A$201",
    'C': "'Products'!$A$2:$A$201",
    'D': "'Employees'!$A$2:$A$201",
}

ws_enr.freeze_panes = 'A2'
write_headers(ws_enr, ENR_HEADERS, FILL_ENR, ENR_REQUIRED)
allocate_data_rows(ws_enr, len(ENR_HEADERS))

for col_letter, nr in ENR_DROPDOWN_NAMED.items():
    add_dropdown(ws_enr, col_letter, nr)
for col_letter, formula in ENR_DROPDOWN_FK.items():
    add_dropdown(ws_enr, col_letter, formula)

sample_enr = ['ENR-2026-0001', 'TVS-CUS-FB-26-0001', 'PRD-CRS-2026-001', 'TVS-EMP-2026-0003', 3500, 'ACTIVE', '2026-03-01', '']
write_sample_row(ws_enr, 2, sample_enr, len(ENR_HEADERS))

col_widths_enr = [20, 24, 22, 22, 14, 14, 14, 28]
for i, w in enumerate(col_widths_enr, start=1):
    ws_enr.column_dimensions[get_column_letter(i)].width = w

# ---------------------------------------------------------------------------
# 8. Ingredients sheet
# ---------------------------------------------------------------------------

ws_ing = wb.create_sheet('Ingredients')
FILL_ING = PatternFill('solid', start_color='33691E', end_color='33691E')

ING_HEADERS = [
    'ingredientId*', 'name*', 'unit*', 'currentStock', 'minStock',
    'category', 'costPerUnit'
]
ING_REQUIRED = {'ingredientId*', 'name*', 'unit*'}
ING_DROPDOWN = {
    'C': 'unit_list',
    'F': 'ingredient_category',
}

ws_ing.freeze_panes = 'A2'
write_headers(ws_ing, ING_HEADERS, FILL_ING, ING_REQUIRED)
allocate_data_rows(ws_ing, len(ING_HEADERS))

for col_letter, nr in ING_DROPDOWN.items():
    add_dropdown(ws_ing, col_letter, nr)

samples_ing = [
    ['ING-2026-001', 'ข้าวญี่ปุ่น', 'กก.', 50, 10, 'grain', 85],
    ['ING-2026-002', 'ปลาทูน่าสด', 'กก.', 5, 2, 'seafood', 450],
    ['ING-2026-003', 'สาหร่ายโนริ', 'แผ่น', 200, 50, 'dry', 8],
]
for i, row_data in enumerate(samples_ing, start=2):
    write_sample_row(ws_ing, i, row_data, len(ING_HEADERS))

col_widths_ing = [18, 22, 10, 14, 12, 18, 14]
for i, w in enumerate(col_widths_ing, start=1):
    ws_ing.column_dimensions[get_column_letter(i)].width = w

# ---------------------------------------------------------------------------
# 9. CourseBOM sheet
# ---------------------------------------------------------------------------

ws_bom = wb.create_sheet('CourseBOM')
FILL_BOM = PatternFill('solid', start_color='006064', end_color='006064')

BOM_HEADERS = ['productId*', 'ingredientId*', 'qtyPerPerson*', 'unit*']
BOM_REQUIRED = {'productId*', 'ingredientId*', 'qtyPerPerson*', 'unit*'}
BOM_DROPDOWN_FK = {
    'A': "'Products'!$A$2:$A$201",
    'B': "'Ingredients'!$A$2:$A$201",
}
BOM_DROPDOWN_NAMED = {
    'D': 'unit_list',
}

ws_bom.freeze_panes = 'A2'
write_headers(ws_bom, BOM_HEADERS, FILL_BOM, BOM_REQUIRED)
allocate_data_rows(ws_bom, len(BOM_HEADERS))

for col_letter, formula in BOM_DROPDOWN_FK.items():
    add_dropdown(ws_bom, col_letter, formula)
for col_letter, nr in BOM_DROPDOWN_NAMED.items():
    add_dropdown(ws_bom, col_letter, nr)

samples_bom = [
    ['PRD-CRS-2026-001', 'ING-2026-001', 0.15, 'กก.'],
    ['PRD-CRS-2026-001', 'ING-2026-002', 0.08, 'กก.'],
    ['PRD-CRS-2026-001', 'ING-2026-003', 3, 'แผ่น'],
]
for i, row_data in enumerate(samples_bom, start=2):
    write_sample_row(ws_bom, i, row_data, len(BOM_HEADERS))

col_widths_bom = [24, 22, 16, 12]
for i, w in enumerate(col_widths_bom, start=1):
    ws_bom.column_dimensions[get_column_letter(i)].width = w

# ---------------------------------------------------------------------------
# 10. Recipes sheet
# ---------------------------------------------------------------------------

ws_rcp = wb.create_sheet('Recipes')
FILL_RCP = PatternFill('solid', start_color='BF360C', end_color='BF360C')

RCP_HEADERS = [
    'recipeId*', 'name*', 'chef', 'category', 'sellingPrice',
    'estimatedCost', 'isActive', 'description'
]
RCP_REQUIRED = {'recipeId*', 'name*'}
RCP_DROPDOWN = {
    'C': 'recipe_chef',
    'D': 'recipe_category',
    'G': 'bool_field',
}

ws_rcp.freeze_panes = 'A2'
write_headers(ws_rcp, RCP_HEADERS, FILL_RCP, RCP_REQUIRED)
allocate_data_rows(ws_rcp, len(RCP_HEADERS))

for col_letter, nr in RCP_DROPDOWN.items():
    add_dropdown(ws_rcp, col_letter, nr)

samples_rcp = [
    ['RCP-AOI-JP-001', 'ซูชิมากุโร่', 'AOI', 'JP', 180, 65, 'TRUE', 'ข้าวซูชิกับมากุโร่สด'],
    ['RCP-FAH-JP-001', 'ราเมนโชยุ', 'FAH', 'JP', 150, 45, 'TRUE', 'ราเมนน้ำใสรสโชยุ'],
]
for i, row_data in enumerate(samples_rcp, start=2):
    write_sample_row(ws_rcp, i, row_data, len(RCP_HEADERS))

col_widths_rcp = [20, 24, 10, 12, 14, 14, 10, 32]
for i, w in enumerate(col_widths_rcp, start=1):
    ws_rcp.column_dimensions[get_column_letter(i)].width = w

# ---------------------------------------------------------------------------
# 11. Packages sheet
# ---------------------------------------------------------------------------

ws_pkg = wb.create_sheet('Packages')
FILL_PKG = PatternFill('solid', start_color='4527A0', end_color='4527A0')

PKG_HEADERS = [
    'packageId*', 'name*', 'originalPrice*', 'packagePrice*',
    'isActive', 'description'
]
PKG_REQUIRED = {'packageId*', 'name*', 'originalPrice*', 'packagePrice*'}
PKG_DROPDOWN = {
    'E': 'bool_field',
}

ws_pkg.freeze_panes = 'A2'
write_headers(ws_pkg, PKG_HEADERS, FILL_PKG, PKG_REQUIRED)
allocate_data_rows(ws_pkg, len(PKG_HEADERS))

for col_letter, nr in PKG_DROPDOWN.items():
    add_dropdown(ws_pkg, col_letter, nr)

sample_pkg = ['PKG-2026-001', 'แพ็กเกจซูชิ+ราเมน', 6300, 5800, 'TRUE', 'เรียน 2 คอร์สพร้อมกัน']
write_sample_row(ws_pkg, 2, sample_pkg, len(PKG_HEADERS))

col_widths_pkg = [18, 28, 16, 16, 10, 36]
for i, w in enumerate(col_widths_pkg, start=1):
    ws_pkg.column_dimensions[get_column_letter(i)].width = w

# ---------------------------------------------------------------------------
# 12. Assets sheet
# ---------------------------------------------------------------------------

ws_ast = wb.create_sheet('Assets')
FILL_AST = PatternFill('solid', start_color='B71C1C', end_color='B71C1C')

AST_HEADERS = [
    'assetId*', 'name*', 'category*', 'status*', 'location',
    'assignedToId', 'purchaseDate', 'purchasePrice', 'vendor',
    'serialNumber', 'warrantyExpiry', 'notes'
]
AST_REQUIRED = {'assetId*', 'name*', 'category*', 'status*'}
AST_DROPDOWN_NAMED = {
    'C': 'asset_category',
    'D': 'asset_status',
}
AST_DROPDOWN_FK = {
    'F': "'Employees'!$A$2:$A$201",
}

ws_ast.freeze_panes = 'A2'
write_headers(ws_ast, AST_HEADERS, FILL_AST, AST_REQUIRED)
allocate_data_rows(ws_ast, len(AST_HEADERS))

for col_letter, nr in AST_DROPDOWN_NAMED.items():
    add_dropdown(ws_ast, col_letter, nr)
for col_letter, formula in AST_DROPDOWN_FK.items():
    add_dropdown(ws_ast, col_letter, formula)

samples_ast = [
    ['AST-KIT-2026-001', 'มีดซูชิมืออาชีพ', 'KITCHEN', 'ACTIVE', 'ครัวห้อง A',
     'TVS-EMP-2026-0004', '2026-01-10', 2500, 'Yoshida Knives', 'YK-2026-001', '2027-01-10', 'คมมาก ระวัง'],
    ['AST-EQP-2026-001', 'โปรเจกเตอร์ Epson', 'EQUIPMENT', 'ACTIVE', 'ห้องเรียน 1',
     '', '2026-01-15', 25000, 'Epson Thailand', 'EPS-X2026', '2028-01-15', ''],
]
for i, row_data in enumerate(samples_ast, start=2):
    write_sample_row(ws_ast, i, row_data, len(AST_HEADERS))

col_widths_ast = [22, 26, 14, 14, 18, 22, 14, 14, 20, 18, 16, 28]
for i, w in enumerate(col_widths_ast, start=1):
    ws_ast.column_dimensions[get_column_letter(i)].width = w

# ---------------------------------------------------------------------------
# 13. Tasks sheet
# ---------------------------------------------------------------------------

ws_tsk = wb.create_sheet('Tasks')
FILL_TSK = PatternFill('solid', start_color='37474F', end_color='37474F')

TSK_HEADERS = [
    'taskId*', 'title*', 'customerId', 'assigneeId', 'type*',
    'priority*', 'status*', 'dueDate', 'description'
]
TSK_REQUIRED = {'taskId*', 'title*', 'type*', 'priority*', 'status*'}
TSK_DROPDOWN_NAMED = {
    'E': 'task_type',
    'F': 'task_priority',
    'G': 'task_status',
}
TSK_DROPDOWN_FK = {
    'C': "'Customers'!$A$2:$A$201",
    'D': "'Employees'!$A$2:$A$201",
}

ws_tsk.freeze_panes = 'A2'
write_headers(ws_tsk, TSK_HEADERS, FILL_TSK, TSK_REQUIRED)
allocate_data_rows(ws_tsk, len(TSK_HEADERS))

for col_letter, nr in TSK_DROPDOWN_NAMED.items():
    add_dropdown(ws_tsk, col_letter, nr)
for col_letter, formula in TSK_DROPDOWN_FK.items():
    add_dropdown(ws_tsk, col_letter, formula)

sample_tsk = [
    'TSK-20260316-001', 'ติดตามลูกค้าหลังทดลองเรียน',
    'TVS-CUS-FB-26-0001', 'TVS-EMP-2026-0003',
    'FOLLOW_UP', 'MEDIUM', 'PENDING', '2026-03-20', 'โทรถามความพึงพอใจ'
]
write_sample_row(ws_tsk, 2, sample_tsk, len(TSK_HEADERS))

col_widths_tsk = [22, 32, 24, 22, 14, 12, 14, 14, 32]
for i, w in enumerate(col_widths_tsk, start=1):
    ws_tsk.column_dimensions[get_column_letter(i)].width = w

# ---------------------------------------------------------------------------
# Reorder sheets: _Enums first, then คู่มือ, then data sheets
# ---------------------------------------------------------------------------
# openpyxl sheet order is determined by wb._sheets list
desired_order = [
    '_Enums', 'คู่มือ', 'Customers', 'Products', 'Employees',
    'CourseSchedules', 'Enrollments', 'Ingredients', 'CourseBOM',
    'Recipes', 'Packages', 'Assets', 'Tasks'
]

# Sort wb._sheets by desired_order
sheet_map = {ws.title: ws for ws in wb._sheets}
wb._sheets = [sheet_map[name] for name in desired_order if name in sheet_map]

# ---------------------------------------------------------------------------
# Save
# ---------------------------------------------------------------------------

wb.save(OUTPUT_PATH)
print(f"Done! Saved to: {OUTPUT_PATH}")
print(f"Sheets: {[ws.title for ws in wb.worksheets]}")
